import openpyxl
import pymysql

conn = pymysql.connect(host='localhost', user='root', passwd='mysql', db='repairtstdb', port=3306, charset='utf8')
cur = conn.cursor()
    
def storeCode(code, name, parentCode, level, provCode, provName):
    sql = "INSERT INTO n_dadi_sec(code, name, parent_code, level, prov_code, prov_name, create_time) VALUES (%s, %s, %s, %s, %s, %s, Now())"
    cur.execute(sql, (code, name, parentCode, level, provCode, provName))
    cur.connection.commit()
    print("****** 区域: %s has been saved to database! ******" % name);

def crawler(startRow, endRow):
    wb = openpyxl.load_workbook('n_dadi_sec.xlsx')
    sheet = wb.get_sheet_by_name('n_dadi_sec')
    
    for i in range(startRow, endRow):
        parentCode = ''
        level = 0
        code = sheet.cell(row=i, column=1).value
        name = sheet.cell(row=i, column=2).value
        code4 = sheet.cell(row=i, column=3).value
        name4 = sheet.cell(row=i, column=4).value
        code3 = sheet.cell(row=i, column=5).value
        name3 = sheet.cell(row=i, column=6).value 
        code2 = sheet.cell(row=i, column=7).value
        name2 = sheet.cell(row=i, column=8).value 
        
        if code4 != code:
            parentCode = code4
            level = 4
        elif code3 != code:
            parentCode = code3
            level = 3
        elif code2 != code:
            parentCode = code2
            level = 2
        else:
            parentCode = ''
            level = 1
        
        provCode = sheet.cell(row=i, column=9).value
        provName = sheet.cell(row=i, column=10).value  
        
        print('code = %s, name = %s, parentCode = %s, level = %d, provCode = %s, provName = %s' % (code, name, parentCode, level, provCode, provName))
        storeCode(code, name, parentCode, level, provCode, provName)

# main
print('****** Crawler start, please wait... ****** ')
crawler(22348, 22349)

print('****** End of crawling ******')